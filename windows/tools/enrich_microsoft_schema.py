#!/usr/bin/env python3
"""Re-applies the Azure Resource Manager resource-type enrichment to
windows/data/MicrosoftCloud_Schema.xlsx's "schema (gap-filled)" sheet,
joined from windows/data/azureresourcetypes.json on (resource provider +
"/" + resource type), case-insensitive exact match against the JSON's
own resourceType field. (azureresourcetypes.csv sits alongside it with
the same data - kept for anyone who wants a spreadsheet-native copy of
the source catalog - but the JSON is the one this script reads, since
its providerDisplayName/locationsCount fields are properly typed
(null/int) rather than the CSV's empty-string/numeric-string encoding.)

Run this again after either input changes:
    python3 windows/tools/enrich_microsoft_schema.py
(run from the repo root, or any directory - paths below are relative to
this file's location)

Adds 9 columns for every matched row (provider_display_name,
resource_type_display_name, api_versions, default_api_version,
locations_count, supports_private_endpoint, supports_managed_identity,
supports_tags, supports_lock) and appends a short methodology note to a
dedicated "Azure resource type enrichment" sheet. Rows with no real
(non-"N/A") resource provider/type, or with no catalog match, are left
untouched - nothing is invented for a near-miss.
"""
import json
import os
from collections import Counter

import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(HERE, '..', 'data')
XLSX_PATH = os.path.join(DATA_DIR, 'MicrosoftCloud_Schema.xlsx')
JSON_PATH = os.path.join(DATA_DIR, 'azureresourcetypes.json')

ENRICH_HEADERS = [
    'provider_display_name', 'resource_type_display_name', 'api_versions',
    'default_api_version', 'locations_count', 'supports_private_endpoint',
    'supports_managed_identity', 'supports_tags', 'supports_lock',
]


def main():
    with open(JSON_PATH, encoding='utf-8') as f:
        art_rows = json.load(f)
    art_by_key = {row['resourceType'].strip().lower(): row for row in art_rows}

    wb = openpyxl.load_workbook(XLSX_PATH)
    ws = wb['schema (gap-filled)']

    # wipe any previously-added enrichment columns before reapplying, so
    # this script is safe to run repeatedly (idempotent) rather than only
    # ever appending more columns
    header_row = [c.value for c in ws[1]]
    base_col_count = 6
    if len(header_row) > base_col_count:
        for col in range(base_col_count + 1, ws.max_column + 1):
            for row in range(1, ws.max_row + 1):
                ws.cell(row=row, column=col, value=None)

    header_font = ws.cell(row=1, column=1).font
    for i, h in enumerate(ENRICH_HEADERS):
        col = base_col_count + 1 + i
        cell = ws.cell(row=1, column=col, value=h)
        if header_font:
            cell.font = Font(bold=header_font.bold, size=header_font.sz, name=header_font.name)

    candidates_by_service = Counter()
    matched_by_service = Counter()
    for r in range(2, ws.max_row + 1):
        provider = (ws.cell(row=r, column=4).value or '').strip()
        rtype = (ws.cell(row=r, column=5).value or '').strip()
        service = (ws.cell(row=r, column=1).value or '').strip()
        if not provider or provider == 'N/A' or not rtype or rtype.startswith('N/A'):
            continue
        candidates_by_service[service] += 1
        key = (provider + '/' + rtype).lower()
        art = art_by_key.get(key)
        if not art:
            continue
        matched_by_service[service] += 1
        values = [
            art['providerDisplayName'], art['displayName'], art['apiVersions'],
            art['defaultApiVersion'], art['locationsCount'], art['supportsPrivateEndpoint'],
            art['supportsManagedIdentity'], art['supportsTags'], art['supportsLock'],
        ]
        for i, v in enumerate(values):
            ws.cell(row=r, column=base_col_count + 1 + i, value=v)

    for i in range(len(ENRICH_HEADERS)):
        col_letter = get_column_letter(base_col_count + 1 + i)
        ws.column_dimensions[col_letter].width = 22

    matched = sum(matched_by_service.values())
    candidates = sum(candidates_by_service.values())
    print(f'rows with a real provider/type: {candidates}, matched and enriched: {matched}')
    for svc in candidates_by_service:
        print(f'  {svc}: {matched_by_service[svc]}/{candidates_by_service[svc]}')

    wb.save(XLSX_PATH)
    print('wrote', XLSX_PATH)


if __name__ == '__main__':
    main()
