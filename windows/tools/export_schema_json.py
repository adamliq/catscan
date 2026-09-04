#!/usr/bin/env python3
"""Exports MicrosoftCloud_Schema.xlsx's "schema (gap-filled)" sheet (the
v20 export, enriched with ARM resource-type metadata) as JSON, matching
the clean key style already used by data/cloud_actions.json (service/
category/operation/provider/resource_type/source) and typing the new
enrichment fields properly (api_versions as an array, locations_count as
an integer, the supports_* flags as real booleans).

Also carries the sheet's own "api" column when a row has one (currently
only the 847 purview rows sourced from Microsoft's raw Office 365
Management Activity API schema reference are tagged, with the value
"Office 365 Management Activity" - distinguishing them from purview
rows sourced from workload-research documentation instead) - added as
an optional `api` key, present only when the sheet cell isn't blank,
same convention as `arm`.

Re-derives the ARM match directly from azureresourcetypes.json rather
than reading the xlsx's already-written enrichment columns back and
guessing from whether they're non-empty - some matched rows have a
genuinely null providerDisplayName in the source (293 catalog rows),
which would otherwise look identical to "not matched".

Run this again after re-running enrich_microsoft_schema.py:
    python3 windows/tools/export_schema_json.py
(run from the repo root, or any directory - paths below are relative to
this file's location)"""
import json
import os
import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(HERE, '..', 'data')
XLSX_PATH = os.path.join(DATA_DIR, 'MicrosoftCloud_Schema.xlsx')
JSON_SOURCE_PATH = os.path.join(DATA_DIR, 'azureresourcetypes.json')
OUT_PATH = os.path.join(DATA_DIR, 'MicrosoftCloud_Schema.json')


def to_bool(v):
    s = (v or '').strip().lower()
    if s == 'true':
        return True
    if s == 'false':
        return False
    return None  # e.g. "undetermined", or blank


def to_int(v):
    if isinstance(v, int):
        return v
    v = (v or '').strip()
    try:
        return int(v)
    except ValueError:
        return None


def main():
    with open(JSON_SOURCE_PATH, encoding='utf-8') as f:
        art_rows = json.load(f)
    art_by_key = {row['resourceType'].strip().lower(): row for row in art_rows}

    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    ws = wb['schema (gap-filled)']
    header = [c.value for c in ws[1]]
    api_col = header.index('api') + 1 if 'api' in header else None
    rows = list(ws.iter_rows(min_row=2, max_col=6, values_only=True))

    out = []
    matched = 0
    api_tagged = 0
    for i, (service, category, operation, provider, rtype, source) in enumerate(rows):
        provider = (provider or '').strip()
        rtype = (rtype or '').strip()
        entry = {
            'service': (service or '').strip(),
            'category': (category or '').strip(),
            'operation': (operation or '').strip(),
            'provider': provider,
            'resource_type': rtype,
            'source': (source or '').strip(),
        }
        if api_col:
            api_val = ws.cell(row=i + 2, column=api_col).value
            if api_val:
                entry['api'] = api_val.strip()
                api_tagged += 1
        art = None
        if provider and provider != 'N/A' and rtype and not rtype.startswith('N/A'):
            art = art_by_key.get((provider + '/' + rtype).lower())
        if art:
            matched += 1
            entry['arm'] = {
                'provider_display_name': art['providerDisplayName'] or None,
                'resource_type_display_name': art['displayName'] or None,
                'api_versions': [v.strip() for v in art['apiVersions'].split(',') if v.strip()],
                'default_api_version': art['defaultApiVersion'] or None,
                'locations_count': to_int(art['locationsCount']),
                'supports_private_endpoint': to_bool(art['supportsPrivateEndpoint']),
                'supports_managed_identity': to_bool(art['supportsManagedIdentity']),
                'supports_tags': to_bool(art['supportsTags']),
                'supports_lock': to_bool(art['supportsLock']),
            }
        out.append(entry)

    with open(OUT_PATH, 'w', encoding='utf-8') as f:
        json.dump(out, f, ensure_ascii=False, indent=2)
        f.write('\n')

    print(f'wrote {OUT_PATH}: {len(out)} rows, {matched} with ARM enrichment, {api_tagged} with an api tag')


if __name__ == '__main__':
    main()
